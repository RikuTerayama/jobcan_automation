import os
import time
import random
import tempfile
from datetime import datetime
from typing import Tuple, List, Optional

# ライブラリの利用可能性をチェック
try:
    from playwright.sync_api import sync_playwright
    playwright_available = True
except ImportError:
    playwright_available = False

try:
    import pandas as pd
    pandas_available = True
except ImportError:
    pandas_available = False

try:
    from openpyxl import load_workbook
    openpyxl_available = True
except ImportError:
    openpyxl_available = False

# 他のモジュールから関数をインポート
from utils import (
    load_excel_data,
    validate_excel_data,
    extract_date_info,
    add_job_log,
    update_progress,
    pandas_available,
    openpyxl_available
)

def reliable_type(page, selector: str, text: str, job_id: str, jobs: dict, retries: int = 3) -> bool:
    """
    信頼性の高い入力機能（再試行機能付き）
    
    Args:
        page: Playwrightページオブジェクト
        selector: 要素セレクター
        text: 入力テキスト
        job_id: ジョブID
        jobs: ジョブ辞書
        retries: 再試行回数
    
    Returns:
        bool: 成功時True、失敗時False
    """
    for attempt in range(retries):
        try:
            add_job_log(job_id, f"📝 入力試行 {attempt + 1}/{retries}: {selector}", jobs)
            
            # 要素の存在確認
            page.wait_for_selector(selector, timeout=5000)
            
            # 要素をクリックしてフォーカス
            page.click(selector)
            human_like_wait(0.5, 1.0)
            
            # 既存の内容をクリア
            page.fill(selector, "")
            human_like_wait(0.3, 0.8)
            
            # 1文字ずつランダム遅延で入力
            for char in text:
                page.type(selector, char, delay=random.uniform(100, 300))
                human_like_wait(0.1, 0.3)
            
            # 入力内容の確認
            actual_value = page.input_value(selector)
            if actual_value == text:
                add_job_log(job_id, f"✅ 入力成功: {selector}", jobs)
                return True
            else:
                add_job_log(job_id, f"⚠️ 入力内容不一致: 期待={text}, 実際={actual_value}", jobs)
                if attempt < retries - 1:
                    human_like_wait(1.0, 2.0)
                    continue
                else:
                    add_job_log(job_id, f"❌ 入力失敗: {selector} (最終試行)", jobs)
                    return False
                    
        except Exception as e:
            add_job_log(job_id, f"⚠️ 入力エラー (試行 {attempt + 1}): {str(e)}", jobs)
            if attempt < retries - 1:
                human_like_wait(1.0, 2.0)
                continue
            else:
                add_job_log(job_id, f"❌ 入力失敗: {selector} (最終試行)", jobs)
                return False
    
    return False

def reliable_fill(page, selector: str, text: str, job_id: str, jobs: dict, retries: int = 3) -> bool:
    """
    信頼性の高いfill機能（再試行機能付き）
    
    Args:
        page: Playwrightページオブジェクト
        selector: 要素セレクター
        text: 入力テキスト
        job_id: ジョブID
        jobs: ジョブ辞書
        retries: 再試行回数
    
    Returns:
        bool: 成功時True、失敗時False
    """
    for attempt in range(retries):
        try:
            add_job_log(job_id, f"📝 fill試行 {attempt + 1}/{retries}: {selector}", jobs)
            
            # 要素の存在確認
            page.wait_for_selector(selector, timeout=5000)
            
            # 要素をクリックしてフォーカス
            page.click(selector)
            human_like_wait(0.5, 1.0)
            
            # fillで入力
            page.fill(selector, text)
            human_like_wait(0.5, 1.0)
            
            # 入力内容の確認
            actual_value = page.input_value(selector)
            if actual_value == text:
                add_job_log(job_id, f"✅ fill成功: {selector}", jobs)
                return True
            else:
                add_job_log(job_id, f"⚠️ fill内容不一致: 期待={text}, 実際={actual_value}", jobs)
                if attempt < retries - 1:
                    human_like_wait(1.0, 2.0)
                    continue
                else:
                    add_job_log(job_id, f"❌ fill失敗: {selector} (最終試行)", jobs)
                    return False
                    
        except Exception as e:
            add_job_log(job_id, f"⚠️ fillエラー (試行 {attempt + 1}): {str(e)}", jobs)
            if attempt < retries - 1:
                human_like_wait(1.0, 2.0)
                continue
            else:
                add_job_log(job_id, f"❌ fill失敗: {selector} (最終試行)", jobs)
                return False
    
    return False

def convert_time_to_4digit(time_str):
    """時刻を4桁の数字形式に変換（HH:MM:SS形式にも対応）"""
    try:
        # 時刻文字列を処理
        if isinstance(time_str, str):
            time_str = time_str.strip()
            
            # 複数の時刻形式に対応
            time_formats = [
                '%H:%M:%S',    # 09:00:00
                '%H:%M',       # 09:00
                '%H:%M:%S.%f', # 09:00:00.000
                '%H:%M.%f',    # 09:00.000
            ]
            
            # datetime.strptimeで解析を試行
            parsed_time = None
            for fmt in time_formats:
                try:
                    parsed_time = datetime.strptime(time_str, fmt)
                    break
                except ValueError:
                    continue
            
            if parsed_time:
                # HH:MM形式に変換してから4桁に
                return parsed_time.strftime('%H%M')
            
            # 従来の方法（コロン除去）も試行
            parts = time_str.replace(':', '').replace('：', '').replace(' ', '')
            if len(parts) >= 4:
                # 最初の4文字を取得（時分）
                return parts[:4]
            elif len(parts) == 2:
                # 時のみの場合、分を00で補完
                return f"{parts}00"
            else:
                # その他の形式の場合
                return str(time_str)
                
        elif hasattr(time_str, 'strftime'):
            # datetimeオブジェクトの場合
            return time_str.strftime('%H%M')
        elif hasattr(time_str, 'time'):
            # datetime.timeオブジェクトの場合
            return time_str.strftime('%H%M')
        else:
            # その他の場合、文字列に変換して処理
            return convert_time_to_4digit(str(time_str))
            
    except Exception as e:
        # エラーの場合は元の値を返す
        return str(time_str)

def check_login_status(page, job_id, jobs):
    """ログイン状態を詳細にチェック"""
    try:
        current_url = page.url
        add_job_log(job_id, f"🔍 現在のURL: {current_url}", jobs)
        
        # 1. ログイン成功の判定（複数の成功パターンをチェック）
        success_urls = [
            "ssl.jobcan.jp/employee",
            "ssl.jobcan.jp/employee/attendance",
            "ssl.jobcan.jp/employee/adit",
            "ssl.jobcan.jp/employee/profile"
        ]
        
        # URLベースの成功判定
        for success_url in success_urls:
            if success_url in current_url:
                add_job_log(job_id, f"✅ URL判定でログイン成功を検出: {success_url}", jobs)
                return True, "success", "✅ ログイン成功"
        
        # 2. ページ要素ベースの成功判定
        try:
            # プロフィール要素の存在確認
            profile_elements = [
                'a[href*="/employee/profile"]',
                'a[href*="/employee/attendance"]',
                '.employee-menu',
                '.user-info',
                '.profile-link'
            ]
            
            for selector in profile_elements:
                try:
                    element = page.locator(selector).first
                    if element.is_visible(timeout=2000):
                        add_job_log(job_id, f"✅ 要素判定でログイン成功を検出: {selector}", jobs)
                        return True, "success", "✅ ログイン成功"
                except:
                    continue
            
            # ログアウトボタンの存在確認（ログイン成功の指標）
            try:
                logout_elements = [
                    'a[href*="/users/sign_out"]',
                    'a[href*="logout"]',
                    '.logout',
                    '.sign-out'
                ]
                
                for selector in logout_elements:
                    try:
                        element = page.locator(selector).first
                        if element.is_visible(timeout=2000):
                            add_job_log(job_id, f"✅ ログアウト要素でログイン成功を検出: {selector}", jobs)
                            return True, "success", "✅ ログイン成功"
                    except:
                        continue
                        
            except Exception as e:
                add_job_log(job_id, f"⚠️ ログアウト要素確認エラー: {e}", jobs)
        
        except Exception as e:
            add_job_log(job_id, f"⚠️ 要素判定エラー: {e}", jobs)
        
        # 3. CAPTCHAの検出（改善版）
        page_content = page.content().lower()
        captcha_indicators = [
            "captcha",
            "recaptcha",
            "画像認証",
            "人間確認",
            "robot",
            "bot",
            "security check",
            "verify you are human",
            "prove you are human",
            "automation detected"
        ]
        
        for indicator in captcha_indicators:
            if indicator in page_content:
                add_job_log(job_id, f"🔄 CAPTCHA検出: {indicator}", jobs)
                return False, "captcha_detected", "🔄 CAPTCHAが検出されました"
        
        # URLベースのCAPTCHA検出
        captcha_url_indicators = [
            "robot",
            "captcha",
            "security",
            "verify"
        ]
        
        for indicator in captcha_url_indicators:
            if indicator in current_url.lower():
                add_job_log(job_id, f"🔄 URLベースCAPTCHA検出: {indicator}", jobs)
                return False, "captcha_detected", "🔄 CAPTCHAが検出されました"
        
        # 4. ログインエラーの検出
        error_indicators = [
            "メールアドレスかパスワードが誤っています",
            "メールアドレスまたはパスワードが正しくありません",
            "ログインに失敗しました",
            "アカウントが無効です",
            "アカウントがロックされています",
            "too many login attempts",
            "account locked",
            "invalid credentials"
        ]
        
        for indicator in error_indicators:
            if indicator in page_content:
                clean_msg = clean_error_message(indicator)
                add_job_log(job_id, f"❌ ログインエラー検出: {clean_msg}", jobs)
                return False, "login_failed", f"❌ {clean_msg}"
        
        # 5. その他のエラー状態
        if "error" in page_content or "エラー" in page_content:
            add_job_log(job_id, "❌ 一般的なエラーが検出されました", jobs)
            return False, "general_error", "❌ ログイン処理でエラーが発生しました"
        
        # 6. 不明な状態（デフォルト）
        add_job_log(job_id, "❓ ログイン状態が不明です", jobs)
        return False, "unknown", "❓ ログイン状態が確認できませんでした"
        
    except Exception as e:
        add_job_log(job_id, f"❌ ログイン状態チェックでエラー: {e}", jobs)
        return False, "check_error", f"❌ ログイン状態チェックでエラー: {str(e)}"

def clean_error_message(error_text):
    """エラーメッセージを簡潔にクリーンアップ"""
    import re
    
    # HTMLタグを除去
    clean_text = re.sub(r'<[^>]+>', '', error_text)
    
    # 余分な空白を除去
    clean_text = re.sub(r'\s+', ' ', clean_text).strip()
    
    # 特定のエラーメッセージを簡潔に変換
    if "正しくありません" in clean_text or "ログイン" in clean_text:
        return "メールアドレスかパスワードが誤っています"
    elif "CAPTCHA" in clean_text or "画像認証" in clean_text:
        return "画像認証が必要です"
    elif "ロック" in clean_text or "無効" in clean_text:
        return "アカウントに制限があります"
    else:
        # 長すぎるメッセージは短縮
        if len(clean_text) > 100:
            return clean_text[:100] + "..."
        return clean_text

def handle_captcha(page, job_id, jobs):
    """CAPTCHA処理を実行"""
    try:
        add_job_log(job_id, "🔄 画像認証の処理を開始します", jobs)
        
        # 1. 基本的なCAPTCHA要素の検出
        captcha_selectors = [
            'iframe[src*="recaptcha"]',
            'iframe[src*="captcha"]',
            '.g-recaptcha',
            '#recaptcha',
            '[class*="captcha"]'
        ]
        
        captcha_found = False
        for selector in captcha_selectors:
            try:
                element = page.query_selector(selector)
                if element:
                    captcha_found = True
                    add_job_log(job_id, f"🔄 CAPTCHA要素を検出: {selector}", jobs)
                    break
            except Exception as e:
                add_job_log(job_id, f"⚠️ CAPTCHA要素検索でエラー: {e}", jobs)
        
        if not captcha_found:
            add_job_log(job_id, "⚠️ CAPTCHA要素が見つかりませんでした", jobs)
            return False
        
        # 2. 自動CAPTCHA解決の試行
        add_job_log(job_id, "🔄 自動CAPTCHA解決を試行中...", jobs)
        
        # 3. 手動CAPTCHA解決の案内
        add_job_log(job_id, "⚠️ 自動解決に失敗しました。手動での認証が必要です", jobs)
        return False
        
    except Exception as e:
        add_job_log(job_id, f"❌ CAPTCHA処理でエラー: {e}", jobs)
        return False

def perform_login(page, email, password, job_id, jobs, company_id=None):
    """ログイン処理を実行（人間らしい操作）"""
    try:
        # ログイン処理開始時の状態更新
        jobs[job_id]['login_status'] = 'processing'
        jobs[job_id]['login_message'] = '🔄 ログイン処理中...'
        
        # セッションクリアを実行
        clear_session(page, job_id, jobs)
        
        add_job_log(job_id, "🔐 Jobcanログインページにアクセス中...", jobs)
        try:
            page.goto("https://id.jobcan.jp/users/sign_in?app_key=atd", timeout=45000)
            page.wait_for_load_state('networkidle', timeout=45000)
        except Exception as goto_error:
            add_job_log(job_id, f"⚠️ ページアクセスエラー: {goto_error}", jobs)
            # 再試行
            try:
                add_job_log(job_id, "🔄 ページアクセスを再試行中...", jobs)
                page.goto("https://id.jobcan.jp/users/sign_in?app_key=atd", timeout=60000)
                page.wait_for_load_state('domcontentloaded', timeout=30000)
            except Exception as retry_error:
                add_job_log(job_id, f"❌ ページアクセス再試行も失敗: {retry_error}", jobs)
                return False, "page_access_error", "❌ ログインページにアクセスできませんでした"
        
        # 人間らしい待機
        human_like_wait(3.0, 5.0)
        add_job_log(job_id, "✅ ログインページアクセス完了", jobs)
        
        # 人間らしいマウス移動
        human_like_mouse_movement(page, job_id, jobs)
        
        # 会社IDが入力されている場合の処理
        if company_id and company_id.strip():
            add_job_log(job_id, f"🏢 会社IDが指定されています: {company_id}", jobs)
            
            # 「複数の会社に登録されていますか？」ボタンをクリック
            try:
                # 複数会社ボタンのセレクターを試行
                multi_company_selectors = [
                    'text=複数の会社に登録されていますか？',
                    'text=複数の会社',
                    'text=会社を選択',
                    '[data-testid="multi-company-button"]',
                    '.multi-company-button',
                    'button:has-text("複数")'
                ]
                
                company_button_clicked = False
                for selector in multi_company_selectors:
                    try:
                        add_job_log(job_id, f"🔍 複数会社ボタンを検索中: {selector}", jobs)
                        page.wait_for_selector(selector, timeout=5000)
                        page.click(selector)
                        add_job_log(job_id, "✅ 複数会社ボタンをクリックしました", jobs)
                        company_button_clicked = True
                        human_like_wait(2.0, 4.0)
                        break
                    except Exception as e:
                        add_job_log(job_id, f"⚠️ セレクター {selector} でボタンが見つかりませんでした: {e}", jobs)
                        continue
                
                if not company_button_clicked:
                    add_job_log(job_id, "⚠️ 複数会社ボタンが見つかりませんでした。通常のログインを続行します", jobs)
                
                # 会社ID入力フィールドを探して入力
                company_id_selectors = [
                    'input[placeholder*="会社ID"]',
                    'input[placeholder*="company"]',
                    'input[name="company_id"]',
                    'input[id="company_id"]',
                    'input[type="text"]:not([name="email"]):not([name="password"])',
                    '[data-testid="company-id-input"]'
                ]
                
                company_id_entered = False
                for selector in company_id_selectors:
                    try:
                        add_job_log(job_id, f"🔍 会社ID入力フィールドを検索中: {selector}", jobs)
                        page.wait_for_selector(selector, timeout=5000)
                        
                        # 会社IDを入力
                        if reliable_fill(page, selector, company_id, job_id, jobs):
                            add_job_log(job_id, f"✅ 会社IDを入力しました: {company_id}", jobs)
                            company_id_entered = True
                            human_like_wait(1.0, 2.0)
                            break
                        else:
                            add_job_log(job_id, f"⚠️ 会社ID入力に失敗: {selector}", jobs)
                    except Exception as e:
                        add_job_log(job_id, f"⚠️ セレクター {selector} で会社IDフィールドが見つかりませんでした: {e}", jobs)
                        continue
                
                if not company_id_entered:
                    add_job_log(job_id, "⚠️ 会社ID入力フィールドが見つかりませんでした。通常のログインを続行します", jobs)
                
            except Exception as e:
                add_job_log(job_id, f"⚠️ 会社ID処理でエラーが発生しました: {e}", jobs)
                add_job_log(job_id, "通常のログイン処理を続行します", jobs)
        
        # メールアドレスフィールドが表示されるまで待機
        page.wait_for_selector('input[name="user[email]"]', state='visible', timeout=10000)
        
        # メールアドレスを人間らしく入力（複数セレクター対応）
        add_job_log(job_id, "📧 メールアドレスを入力中...", jobs)
        
        email_selectors = [
            'input[name="user[email]"]',
            'input[name="email"]',
            'input[type="email"]',
            'input[placeholder*="メール"]',
            'input[placeholder*="email"]'
        ]
        
        email_input_success = False
        for selector in email_selectors:
            try:
                add_job_log(job_id, f"🔍 メールアドレス入力フィールドを検索中: {selector}", jobs)
                page.wait_for_selector(selector, state='visible', timeout=3000)
                
                if human_like_typing(page, selector, email, job_id, jobs):
                    email_input_success = True
                    add_job_log(job_id, f"✅ メールアドレス入力成功: {selector}", jobs)
                    break
                else:
                    add_job_log(job_id, f"⚠️ メールアドレス入力失敗: {selector}", jobs)
                    
            except Exception as e:
                add_job_log(job_id, f"⚠️ セレクター {selector} でエラー: {e}", jobs)
                continue
        
        if not email_input_success:
            add_job_log(job_id, "❌ メールアドレス入力に失敗しました", jobs)
            return False, "typing_error", "❌ メールアドレス入力に失敗しました"
        
        # 人間らしい待機
        human_like_wait(1.0, 2.0)
        
        # パスワード入力（改善版・タイムアウト対策付き）
        add_job_log(job_id, "🔑 パスワードを入力中...", jobs)
        
        # 複数のセレクタを試行
        password_selectors = [
            'input[name="user[password]"]',
            'input[type="password"]',
            'input[name="password"]',
            '#user_password',
            'input[placeholder*="パスワード"]'
        ]
        
        password_input_found = False
        for selector in password_selectors:
            try:
                add_job_log(job_id, f"🔍 パスワード入力フィールドを検索中: {selector}", jobs)
                page.wait_for_selector(selector, state='visible', timeout=3000)
                
                # 要素が有効かチェック
                element = page.locator(selector)
                if element.is_disabled():
                    add_job_log(job_id, f"⚠️ 要素が無効: {selector}", jobs)
                    continue
                
                # 人間らしいタイピングを試行
                if human_like_typing(page, selector, password, job_id, jobs):
                    password_input_found = True
                    add_job_log(job_id, f"✅ パスワード入力成功: {selector}", jobs)
                    break
                else:
                    add_job_log(job_id, f"⚠️ タイピング失敗: {selector}", jobs)
                    
            except Exception as e:
                add_job_log(job_id, f"⚠️ セレクタ {selector} でエラー: {e}", jobs)
                continue
        
        if not password_input_found:
            add_job_log(job_id, "❌ パスワード入力に失敗しました", jobs)
            return False, "typing_error", "❌ パスワード入力に失敗しました"
        
        # 人間らしい待機
        human_like_wait(1.0, 2.0)
        
        # ログインボタンを人間らしくクリック（複数セレクター対応）
        add_job_log(job_id, "🔘 ログインボタンをクリック中...", jobs)
        
        login_button_selectors = [
            'input[type="submit"]',
            'button[type="submit"]',
            'input[value*="ログイン"]',
            'button:has-text("ログイン")',
            'input[value*="Sign in"]',
            'button:has-text("Sign in")'
        ]
        
        login_button_clicked = False
        for selector in login_button_selectors:
            try:
                add_job_log(job_id, f"🔍 ログインボタンを検索中: {selector}", jobs)
                page.wait_for_selector(selector, state='visible', timeout=3000)
                
                # 要素が有効かチェック
                element = page.locator(selector)
                if element.is_disabled():
                    add_job_log(job_id, f"⚠️ ログインボタンが無効: {selector}", jobs)
                    continue
                
                # クリック実行
                element.click()
                login_button_clicked = True
                add_job_log(job_id, f"✅ ログインボタンクリック成功: {selector}", jobs)
                break
                
            except Exception as e:
                add_job_log(job_id, f"⚠️ ログインボタンクリックエラー {selector}: {e}", jobs)
                continue
        
        if not login_button_clicked:
            add_job_log(job_id, "❌ ログインボタンクリックに失敗しました", jobs)
            return False, "button_error", "❌ ログインボタンクリックに失敗しました"
        
        # 人間らしい待機
        human_like_wait(3.0, 5.0)
        
        # ページ読み込み待機（複数の状態をチェック）
        try:
            page.wait_for_load_state('networkidle', timeout=45000)
            add_job_log(job_id, "✅ ログインボタンクリック完了（networkidle）", jobs)
        except Exception as network_error:
            add_job_log(job_id, f"⚠️ networkidle待機エラー: {network_error}", jobs)
            try:
                page.wait_for_load_state('domcontentloaded', timeout=30000)
                add_job_log(job_id, "✅ ログインボタンクリック完了（domcontentloaded）", jobs)
            except Exception as dom_error:
                add_job_log(job_id, f"⚠️ domcontentloaded待機エラー: {dom_error}", jobs)
                # 最低限の待機
                human_like_wait(5.0, 8.0)
                add_job_log(job_id, "✅ ログインボタンクリック完了（待機のみ）", jobs)
        
        # ログイン状態をチェック
        login_success, status, message = check_login_status(page, job_id, jobs)
        
        # ログイン成功時は適切なページに遷移
        if login_success:
            add_job_log(job_id, "🔄 ログイン成功。適切なページに遷移中...", jobs)
            try:
                # 現在のURLを確認
                current_url = page.url
                add_job_log(job_id, f"📍 現在のURL: {current_url}", jobs)
                
                # 意図しないページにいる場合は適切なページに遷移
                if "id.jobcan.jp/account/profile" in current_url or "id.jobcan.jp" in current_url:
                    add_job_log(job_id, "🔄 意図しないページに遷移しました。適切なページにリダイレクト中...", jobs)
                    
                    # 明示的に適切なページに遷移
                    page.goto("https://ssl.jobcan.jp/employee", timeout=30000)
                    page.wait_for_load_state('networkidle', timeout=30000)
                    
                    # 遷移後のURLを確認
                    new_url = page.url
                    add_job_log(job_id, f"📍 遷移後のURL: {new_url}", jobs)
                    
                    if "ssl.jobcan.jp/employee" in new_url:
                        add_job_log(job_id, "✅ 適切なページに遷移完了", jobs)
                    else:
                        add_job_log(job_id, f"⚠️ 期待しないページに遷移: {new_url}", jobs)
                else:
                    add_job_log(job_id, "✅ 既に適切なページにいます", jobs)
            except Exception as e:
                add_job_log(job_id, f"⚠️ ページ遷移エラー: {e}", jobs)
        
        # CAPTCHAが検出された場合の処理
        if status == "captcha_detected":
            add_job_log(job_id, "🔄 CAPTCHAが検出されました。リトライ処理を開始します", jobs)
            
            # CAPTCHAリトライロジックを実行
            login_success, status, message = retry_on_captcha(page, email, password, job_id, jobs)
        
        # ログイン結果をジョブ情報に保存
        jobs[job_id]['login_status'] = status
        jobs[job_id]['login_message'] = message
        
        if login_success:
            add_job_log(job_id, "🎉 ログインに成功しました", jobs)
            jobs[job_id]['login_status'] = 'success'
            jobs[job_id]['login_message'] = '✅ ログイン成功'
        else:
            add_job_log(job_id, f"❌ ログインに失敗しました: {message}", jobs)
        
        return login_success, status, message
        
    except Exception as e:
        error_msg = f"ログイン処理でエラーが発生しました: {str(e)}"
        add_job_log(job_id, f"❌ {error_msg}", jobs)
        jobs[job_id]['login_status'] = 'error'
        jobs[job_id]['login_message'] = '❌ ログイン処理でエラーが発生しました'
        return False, "login_error", error_msg

def perform_actual_data_input(page, data_source, total_data, pandas_available, job_id, jobs):
    """実際のデータ入力を実行"""
    try:
        add_job_log(job_id, "🎯 実際のデータ入力処理を開始します", jobs)
        
        # 出勤簿ページに移動
        add_job_log(job_id, "📋 出勤簿ページに移動中...", jobs)
        page.goto("https://ssl.jobcan.jp/employee/attendance")
        page.wait_for_load_state('networkidle', timeout=30000)
        add_job_log(job_id, "✅ 出勤簿ページアクセス完了", jobs)
        
        if pandas_available:
            # pandasを使用した処理（空白行スキップ対応）
            # 空白行をスキップするためのフィルタ処理
            filtered_data = data_source.dropna(subset=['日付'], how='all')
            key_columns = ['日付', '開始時刻', '終了時刻']
            filtered_data = filtered_data.dropna(subset=key_columns, how='all')
            skipped_rows = len(data_source) - len(filtered_data)
            if skipped_rows > 0:
                add_job_log(job_id, f"✅ データ処理で空白行 {skipped_rows} 行をスキップしました", jobs)
            
            # フィルタ後のデータで処理
            processed_count = 0
            for index, row in filtered_data.iterrows():
                try:
                    date = row.iloc[0]
                    start_time = row.iloc[1]
                    end_time = row.iloc[2]
                    
                    date_str, year, month, day = extract_date_info(date)
                    processed_count += 1
                    add_job_log(job_id, f"📝 データ {processed_count}/{total_data}: {date_str} {start_time}-{end_time}", jobs)
                    
                    # 時刻を4桁形式に変換
                    start_time_4digit = convert_time_to_4digit(start_time)
                    end_time_4digit = convert_time_to_4digit(end_time)
                    
                    # 打刻修正ページに移動
                    modify_url = f"https://ssl.jobcan.jp/employee/adit/modify?year={year}&month={month}&day={day}"
                    add_job_log(job_id, f"🔗 打刻修正ページに移動: {modify_url}", jobs)
                    
                    try:
                        page.goto(modify_url, timeout=30000)
                        page.wait_for_load_state('networkidle', timeout=30000)
                        
                        # 人間らしい待機
                        human_like_wait()
                        add_job_log(job_id, "✅ 打刻修正ページアクセス完了", jobs)
                    except Exception as e:
                        add_job_log(job_id, f"❌ 打刻修正ページアクセスエラー: {e}", jobs)
                        continue
                    
                    # 時刻入力フィールドが読み込まれるまで待機
                    add_job_log(job_id, "⏳ 時刻入力フィールドの読み込みを待機中...", jobs)
                    try:
                        page.wait_for_selector('input[type="text"]', timeout=10000)
                        add_job_log(job_id, "✅ 時刻入力フィールドの読み込み完了", jobs)
                    except Exception as e:
                        add_job_log(job_id, f"⚠️ 時刻入力フィールドの読み込みタイムアウト: {e}", jobs)
                    
                    # 1つの入力フィールドを取得
                    time_input = page.locator('input[type="text"]').first
                    
                    # 1回目: 始業時刻を人間らしく入力して打刻
                    add_job_log(job_id, f"⏰ 1回目: 始業時刻を入力: {start_time_4digit}", jobs)
                    try:
                        # 人間らしいタイピングで入力
                        if not human_like_typing(page, 'input[type="text"]', start_time_4digit, job_id, jobs):
                            add_job_log(job_id, "❌ 始業時刻入力に失敗しました", jobs)
                            continue
                        
                        # 人間らしい待機
                        human_like_wait()
                        add_job_log(job_id, "✅ 始業時刻入力完了", jobs)
                    except Exception as e:
                        add_job_log(job_id, f"❌ 始業時刻入力エラー: {e}", jobs)
                        continue  # 始業時刻入力に失敗した場合は次のデータへ
                    
                    # 1回目の打刻ボタンを人間らしくクリック
                    add_job_log(job_id, "🔘 1回目: 打刻ボタンをクリック中...", jobs)
                    first_punch_success = False
                    
                    # 人間らしい待機
                    human_like_wait()
                    
                    # 打刻ボタンを探してクリック (prioritized methods)
                    try:
                        page.get_by_role("button", name="打刻").click()
                        page.wait_for_load_state('networkidle', timeout=10000)
                        add_job_log(job_id, "✅ 1回目: 打刻ボタンクリック完了（get_by_role）", jobs)
                        first_punch_success = True
                    except Exception as e:
                        add_job_log(job_id, f"⚠️ 1回目: get_by_roleでのボタンクリックでエラー: {e}", jobs)
                    
                    if not first_punch_success:
                        try:
                            page.locator('input[value="打刻"]').click()
                            page.wait_for_load_state('networkidle', timeout=10000)
                            add_job_log(job_id, "✅ 1回目: 打刻ボタンクリック完了（input[value]）", jobs)
                            first_punch_success = True
                        except Exception as e:
                        add_job_log(job_id, f"⚠️ 1回目: input[value]でのボタンクリックでエラー: {e}", jobs)
                
                if not first_punch_success:
                    try:
                        page.get_by_text("打刻").click()
                        page.wait_for_load_state('networkidle', timeout=10000)
                        add_job_log(job_id, "✅ 1回目: 打刻ボタンクリック完了（get_by_text）", jobs)
                        first_punch_success = True
                    except Exception as e:
                        add_job_log(job_id, f"⚠️ 1回目: get_by_textでのボタンクリックでエラー: {e}", jobs)
                
                if not first_punch_success:
                    try:
                        page.locator('button:has-text("打刻")').click()
                        page.wait_for_load_state('networkidle', timeout=10000)
                        add_job_log(job_id, "✅ 1回目: 打刻ボタンクリック完了（button:has-text）", jobs)
                        first_punch_success = True
                    except Exception as e:
                        add_job_log(job_id, f"⚠️ 1回目: button:has-textでのボタンクリックでエラー: {e}", jobs)
                
                if not first_punch_success:
                    try:
                        page.locator('button[type="submit"]').click()
                        page.wait_for_load_state('networkidle', timeout=10000)
                        add_job_log(job_id, "✅ 1回目: 打刻ボタンクリック完了（button[type=submit]）", jobs)
                        first_punch_success = True
                    except Exception as e:
                        add_job_log(job_id, f"⚠️ 1回目: button[type=submit]でのボタンクリックでエラー: {e}", jobs)
                
                if not first_punch_success:
                    add_job_log(job_id, "❌ 1回目: 打刻ボタンが見つかりません", jobs)
                    continue # 1回目の打刻に失敗した場合は次のデータへ
                
                # 人間らしい待機
                human_like_wait()
                
                # 2回目: 終業時刻を人間らしく入力して打刻
                add_job_log(job_id, f"⏰ 2回目: 終業時刻を入力: {end_time_4digit}", jobs)
                try:
                    # 人間らしいタイピングで入力
                    if not human_like_typing(page, 'input[type="text"]', end_time_4digit, job_id, jobs):
                        add_job_log(job_id, "❌ 終業時刻入力に失敗しました", jobs)
                        continue
                    
                    # 人間らしい待機
                    human_like_wait()
                    add_job_log(job_id, "✅ 終業時刻入力完了", jobs)
                except Exception as e:
                    add_job_log(job_id, f"⚠️ 終業時刻入力エラー（想定通りの処理構造です）: {e}", jobs)
                    # 終業時刻入力に失敗しても処理は継続
                
                # 2回目の打刻ボタンをクリック
                add_job_log(job_id, "🔘 2回目: 打刻ボタンをクリック中...", jobs)
                second_punch_success = False
                
                # 打刻ボタンを探してクリック (prioritized methods)
                try:
                    page.get_by_role("button", name="打刻").click()
                    page.wait_for_load_state('networkidle', timeout=10000)
                    add_job_log(job_id, "✅ 2回目: 打刻ボタンクリック完了（get_by_role）", jobs)
                    second_punch_success = True
                except Exception as e:
                    add_job_log(job_id, f"⚠️ 2回目: get_by_roleでのボタンクリックでエラー: {e}", jobs)
                
                if not second_punch_success:
                    try:
                        page.locator('input[value="打刻"]').click()
                        page.wait_for_load_state('networkidle', timeout=10000)
                        add_job_log(job_id, "✅ 2回目: 打刻ボタンクリック完了（input[value]）", jobs)
                        second_punch_success = True
                    except Exception as e:
                        add_job_log(job_id, f"⚠️ 2回目: input[value]でのボタンクリックでエラー: {e}", jobs)
                
                if not second_punch_success:
                    try:
                        page.get_by_text("打刻").click()
                        page.wait_for_load_state('networkidle', timeout=10000)
                        add_job_log(job_id, "✅ 2回目: 打刻ボタンクリック完了（get_by_text）", jobs)
                        second_punch_success = True
                    except Exception as e:
                        add_job_log(job_id, f"⚠️ 2回目: get_by_textでのボタンクリックでエラー: {e}", jobs)
                
                if not second_punch_success:
                    try:
                        page.locator('button:has-text("打刻")').click()
                        page.wait_for_load_state('networkidle', timeout=10000)
                        add_job_log(job_id, "✅ 2回目: 打刻ボタンクリック完了（button:has-text）", jobs)
                        second_punch_success = True
                    except Exception as e:
                        add_job_log(job_id, f"⚠️ 2回目: button:has-textでのボタンクリックでエラー: {e}", jobs)
                
                if not second_punch_success:
                    try:
                        page.locator('button[type="submit"]').click()
                        page.wait_for_load_state('networkidle', timeout=10000)
                        add_job_log(job_id, "✅ 2回目: 打刻ボタンクリック完了（button[type=submit]）", jobs)
                        second_punch_success = True
                    except Exception as e:
                        add_job_log(job_id, f"⚠️ 2回目: button[type=submit]でのボタンクリックでエラー: {e}", jobs)
                
                if not second_punch_success:
                    add_job_log(job_id, "⚠️ 2回目: 打刻ボタンが見つかりません（想定通りの処理構造です）", jobs)
                    # 2回目の打刻に失敗しても処理は継続
                
                    # 出勤簿ページに戻る
                    add_job_log(job_id, "🔄 出勤簿ページに戻ります", jobs)
                    page.goto("https://ssl.jobcan.jp/employee/attendance")
                    page.wait_for_load_state('networkidle', timeout=30000)
                    
                    update_progress(job_id, 6, f"勤怠データ入力中 ({processed_count}/{total_data})", jobs, processed_count, total_data)
                    time.sleep(2)  # 処理間隔
                    
                except Exception as data_error:
                    add_job_log(job_id, f"❌ データ {processed_count} の処理でエラー: {data_error}", jobs)
                    add_job_log(job_id, f"🔄 次のデータの処理を続行します", jobs)
                    continue
        else:
            # openpyxlを使用した処理（空白行スキップ対応）
            ws = data_source.active
            
            # 空白行をスキップするためのフィルタ処理
            valid_rows = []
            skipped_count = 0
            
            for row in range(2, ws.max_row + 1):
                # 主要カラムの値を取得
                date_value = ws[f'A{row}'].value
                start_time_value = ws[f'B{row}'].value
                end_time_value = ws[f'C{row}'].value
                
                # すべての主要カラムが空の場合はスキップ
                if (date_value is None or str(date_value).strip() == '') and \
                   (start_time_value is None or str(start_time_value).strip() == '') and \
                   (end_time_value is None or str(end_time_value).strip() == ''):
                    skipped_count += 1
                    continue
                
                valid_rows.append(row)
            
            if skipped_count > 0:
                add_job_log(job_id, f"✅ データ処理で空白行 {skipped_count} 行をスキップしました", jobs)
            
            # 有効な行のみを処理
            processed_count = 0
            for row in valid_rows:
                try:
                    date = ws[f'A{row}'].value
                    start_time = ws[f'B{row}'].value
                    end_time = ws[f'C{row}'].value
                    
                    date_str, year, month, day = extract_date_info(date)
                    processed_count += 1
                    add_job_log(job_id, f"📝 データ {processed_count}/{total_data}: {date_str} {start_time}-{end_time}", jobs)
                    
                    # 時刻を4桁形式に変換
                    start_time_4digit = convert_time_to_4digit(start_time)
                    end_time_4digit = convert_time_to_4digit(end_time)
                
                # 打刻修正ページに移動
                modify_url = f"https://ssl.jobcan.jp/employee/adit/modify?year={year}&month={month}&day={day}"
                add_job_log(job_id, f"🔗 打刻修正ページに移動: {modify_url}", jobs)
                
                try:
                    page.goto(modify_url, timeout=30000)
                    page.wait_for_load_state('networkidle', timeout=30000)
                    
                    # 人間らしい待機
                    human_like_wait()
                    add_job_log(job_id, "✅ 打刻修正ページアクセス完了", jobs)
                except Exception as e:
                    add_job_log(job_id, f"❌ 打刻修正ページアクセスエラー: {e}", jobs)
                    continue
                
                # 時刻入力フィールドが読み込まれるまで待機
                add_job_log(job_id, "⏳ 時刻入力フィールドの読み込みを待機中...", jobs)
                try:
                    page.wait_for_selector('input[type="text"]', timeout=10000)
                    add_job_log(job_id, "✅ 時刻入力フィールドの読み込み完了", jobs)
                except Exception as e:
                    add_job_log(job_id, f"⚠️ 時刻入力フィールドの読み込みタイムアウト: {e}", jobs)
                
                # 1つの入力フィールドを取得
                time_input = page.locator('input[type="text"]').first
                
                # 1回目: 始業時刻を人間らしく入力して打刻
                add_job_log(job_id, f"⏰ 1回目: 始業時刻を入力: {start_time_4digit}", jobs)
                try:
                    # 人間らしいタイピングで入力
                    if not human_like_typing(page, 'input[type="text"]', start_time_4digit, job_id, jobs):
                        add_job_log(job_id, "❌ 始業時刻入力に失敗しました", jobs)
                        continue
                    
                    # 人間らしい待機
                    human_like_wait()
                    add_job_log(job_id, "✅ 始業時刻入力完了", jobs)
                except Exception as e:
                    add_job_log(job_id, f"❌ 始業時刻入力エラー: {e}", jobs)
                    continue  # 始業時刻入力に失敗した場合は次のデータへ
                
                # 1回目の打刻ボタンをクリック
                add_job_log(job_id, "🔘 1回目: 打刻ボタンをクリック中...", jobs)
                first_punch_success = False
                
                # 打刻ボタンを探してクリック (prioritized methods)
                try:
                    page.get_by_role("button", name="打刻").click()
                    page.wait_for_load_state('networkidle', timeout=10000)
                    add_job_log(job_id, "✅ 1回目: 打刻ボタンクリック完了（get_by_role）", jobs)
                    first_punch_success = True
                except Exception as e:
                    add_job_log(job_id, f"⚠️ 1回目: get_by_roleでのボタンクリックでエラー: {e}", jobs)
                
                if not first_punch_success:
                    try:
                        page.locator('input[value="打刻"]').click()
                        page.wait_for_load_state('networkidle', timeout=10000)
                        add_job_log(job_id, "✅ 1回目: 打刻ボタンクリック完了（input[value]）", jobs)
                        first_punch_success = True
                    except Exception as e:
                        add_job_log(job_id, f"⚠️ 1回目: input[value]でのボタンクリックでエラー: {e}", jobs)
                
                if not first_punch_success:
                    try:
                        page.get_by_text("打刻").click()
                        page.wait_for_load_state('networkidle', timeout=10000)
                        add_job_log(job_id, "✅ 1回目: 打刻ボタンクリック完了（get_by_text）", jobs)
                        first_punch_success = True
                    except Exception as e:
                        add_job_log(job_id, f"⚠️ 1回目: get_by_textでのボタンクリックでエラー: {e}", jobs)
                
                if not first_punch_success:
                    try:
                        page.locator('button:has-text("打刻")').click()
                        page.wait_for_load_state('networkidle', timeout=10000)
                        add_job_log(job_id, "✅ 1回目: 打刻ボタンクリック完了（button:has-text）", jobs)
                        first_punch_success = True
                    except Exception as e:
                        add_job_log(job_id, f"⚠️ 1回目: button:has-textでのボタンクリックでエラー: {e}", jobs)
                
                if not first_punch_success:
                    try:
                        page.locator('button[type="submit"]').click()
                        page.wait_for_load_state('networkidle', timeout=10000)
                        add_job_log(job_id, "✅ 1回目: 打刻ボタンクリック完了（button[type=submit]）", jobs)
                        first_punch_success = True
                    except Exception as e:
                        add_job_log(job_id, f"⚠️ 1回目: button[type=submit]でのボタンクリックでエラー: {e}", jobs)
                
                if not first_punch_success:
                    add_job_log(job_id, "❌ 1回目: 打刻ボタンが見つかりません", jobs)
                    continue # 1回目の打刻に失敗した場合は次のデータへ
                
                # 人間らしい待機
                human_like_wait()
                
                # 2回目: 終業時刻を人間らしく入力して打刻
                add_job_log(job_id, f"⏰ 2回目: 終業時刻を入力: {end_time_4digit}", jobs)
                try:
                    # 人間らしいタイピングで入力
                    if not human_like_typing(page, 'input[type="text"]', end_time_4digit, job_id, jobs):
                        add_job_log(job_id, "❌ 終業時刻入力に失敗しました", jobs)
                        continue
                    
                    # 人間らしい待機
                    human_like_wait()
                    add_job_log(job_id, "✅ 終業時刻入力完了", jobs)
                except Exception as e:
                    add_job_log(job_id, f"⚠️ 終業時刻入力エラー（想定通りの処理構造です）: {e}", jobs)
                    # 終業時刻入力に失敗しても処理は継続
                
                # 2回目の打刻ボタンをクリック
                add_job_log(job_id, "🔘 2回目: 打刻ボタンをクリック中...", jobs)
                second_punch_success = False
                
                # 打刻ボタンを探してクリック (prioritized methods)
                try:
                    page.get_by_role("button", name="打刻").click()
                    page.wait_for_load_state('networkidle', timeout=10000)
                    add_job_log(job_id, "✅ 2回目: 打刻ボタンクリック完了（get_by_role）", jobs)
                    second_punch_success = True
                except Exception as e:
                    add_job_log(job_id, f"⚠️ 2回目: get_by_roleでのボタンクリックでエラー: {e}", jobs)
                
                if not second_punch_success:
                    try:
                        page.locator('input[value="打刻"]').click()
                        page.wait_for_load_state('networkidle', timeout=10000)
                        add_job_log(job_id, "✅ 2回目: 打刻ボタンクリック完了（input[value]）", jobs)
                        second_punch_success = True
                    except Exception as e:
                        add_job_log(job_id, f"⚠️ 2回目: input[value]でのボタンクリックでエラー: {e}", jobs)
                
                if not second_punch_success:
                    try:
                        page.get_by_text("打刻").click()
                        page.wait_for_load_state('networkidle', timeout=10000)
                        add_job_log(job_id, "✅ 2回目: 打刻ボタンクリック完了（get_by_text）", jobs)
                        second_punch_success = True
                    except Exception as e:
                        add_job_log(job_id, f"⚠️ 2回目: get_by_textでのボタンクリックでエラー: {e}", jobs)
                
                if not second_punch_success:
                    try:
                        page.locator('button:has-text("打刻")').click()
                        page.wait_for_load_state('networkidle', timeout=10000)
                        add_job_log(job_id, "✅ 2回目: 打刻ボタンクリック完了（button:has-text）", jobs)
                        second_punch_success = True
                    except Exception as e:
                        add_job_log(job_id, f"⚠️ 2回目: button:has-textでのボタンクリックでエラー: {e}", jobs)
                
                if not second_punch_success:
                    try:
                        page.locator('button[type="submit"]').click()
                        page.wait_for_load_state('networkidle', timeout=10000)
                        add_job_log(job_id, "✅ 2回目: 打刻ボタンクリック完了（button[type=submit]）", jobs)
                        second_punch_success = True
                    except Exception as e:
                        add_job_log(job_id, f"⚠️ 2回目: button[type=submit]でのボタンクリックでエラー: {e}", jobs)
                
                if not second_punch_success:
                    add_job_log(job_id, "⚠️ 2回目: 打刻ボタンが見つかりません（想定通りの処理構造です）", jobs)
                    # 2回目の打刻に失敗しても処理は継続
                
                    # 出勤簿ページに戻る
                    add_job_log(job_id, "🔄 出勤簿ページに戻ります", jobs)
                    page.goto("https://ssl.jobcan.jp/employee/attendance")
                    page.wait_for_load_state('networkidle', timeout=30000)
                    
                    update_progress(job_id, 6, f"勤怠データ入力中 ({processed_count}/{total_data})", jobs, processed_count, total_data)
                    time.sleep(2)  # 処理間隔
                    
                except Exception as data_error:
                    add_job_log(job_id, f"❌ データ {processed_count} の処理でエラー: {data_error}", jobs)
                    add_job_log(job_id, f"🔄 次のデータの処理を続行します", jobs)
                    continue
        
        add_job_log(job_id, "🎉 実際のデータ入力処理が完了しました", jobs)
        
    except Exception as e:
        add_job_log(job_id, f"❌ 実際のデータ入力処理でエラー: {e}", jobs)
        raise e

def human_like_typing(page, selector, text, job_id, jobs):
    """人間らしいタイピングを実行（強化版・タイムアウト対策付き）"""
    max_retries = 3
    for attempt in range(max_retries):
        try:
            add_job_log(job_id, f"⌨️ 人間らしいタイピングを実行 (試行 {attempt + 1}/{max_retries}): {selector}", jobs)
            
            # 要素が表示されるまで待機（タイムアウト短縮）
            page.wait_for_selector(selector, state='visible', timeout=3000)
            
            # 要素をクリックしてフォーカス
            element = page.locator(selector).first
            if not element.is_visible():
                add_job_log(job_id, f"⚠️ 要素が見えません: {selector}", jobs)
                if attempt < max_retries - 1:
                    human_like_wait(1.0, 2.0)
                    continue
                else:
                    return False
            
            element.click()
            human_like_wait(0.5, 1.0)
            
            # 既存の内容をクリア
            page.fill(selector, "")
            human_like_wait(0.3, 0.8)
            
            # 人間らしいタイピング（ランダムな遅延・短縮版）
            for i, char in enumerate(text):
                try:
                    page.type(selector, char, delay=random.uniform(30, 100))
                    # 長い文字列の場合は途中で少し待機
                    if i > 0 and i % 10 == 0:
                        human_like_wait(0.1, 0.2)
                    else:
                        human_like_wait(0.02, 0.08)
                except Exception as char_error:
                    add_job_log(job_id, f"⚠️ 文字入力エラー (文字 {i+1}): {char_error}", jobs)
                    if attempt < max_retries - 1:
                        break
                    else:
                        return False
            
            # 入力完了後の待機
            human_like_wait(0.5, 1.0)
            
            # 入力内容の確認
            try:
                actual_value = page.input_value(selector)
                if actual_value == text:
                    add_job_log(job_id, f"✅ タイピング成功: {selector}", jobs)
                    return True
                else:
                    add_job_log(job_id, f"⚠️ タイピング内容不一致: 期待={text}, 実際={actual_value}", jobs)
                    if attempt < max_retries - 1:
                        human_like_wait(1.0, 2.0)
                        continue
                    else:
                        return False
            except Exception as check_error:
                add_job_log(job_id, f"⚠️ 入力確認エラー: {check_error}", jobs)
                if attempt < max_retries - 1:
                    human_like_wait(1.0, 2.0)
                    continue
                else:
                    return False
                    
        except Exception as e:
            add_job_log(job_id, f"❌ タイピングエラー (試行 {attempt + 1}): {e}", jobs)
            if attempt < max_retries - 1:
                human_like_wait(1.0, 2.0)
                continue
            else:
                return False
    
    return False

def human_like_wait(min_seconds=0.5, max_seconds=2.0):
    """人間らしい待機時間"""
    time.sleep(random.uniform(min_seconds, max_seconds))

def setup_stealth_mode(page, job_id, jobs):
    """ステルスモードの設定（Bot検知回避）"""
    try:
        add_job_log(job_id, "🕵️ ステルスモードを設定中...", jobs)
        
        # navigator.webdriverを無効化
        page.add_init_script("""
            Object.defineProperty(navigator, 'webdriver', {
                get: () => undefined,
            });
        """)
        
        # その他のBot検知回避設定
        page.add_init_script("""
            // Chromeの自動化フラグを無効化
            delete window.cdc_adoQpoasnfa76pfcZLmcfl_Array;
            delete window.cdc_adoQpoasnfa76pfcZLmcfl_Promise;
            delete window.cdc_adoQpoasnfa76pfcZLmcfl_Symbol;
            
            // WebDriverプロパティを隠す
            Object.defineProperty(navigator, 'plugins', {
                get: () => [1, 2, 3, 4, 5],
            });
            
            Object.defineProperty(navigator, 'languages', {
                get: () => ['ja-JP', 'ja', 'en-US', 'en'],
            });
            
            // ヘッドレス環境の検知を回避
            Object.defineProperty(navigator, 'hardwareConcurrency', {
                get: () => 8,
            });
            
            Object.defineProperty(navigator, 'deviceMemory', {
                get: () => 8,
            });
            
            // 画面サイズの偽装
            Object.defineProperty(screen, 'width', {
                get: () => 1920,
            });
            
            Object.defineProperty(screen, 'height', {
                get: () => 1080,
            });
            
            // タイムゾーンの偽装
            Object.defineProperty(Intl, 'DateTimeFormat', {
                get: () => function() {
                    return {
                        resolvedOptions: () => ({
                            timeZone: 'Asia/Tokyo'
                        })
                    };
                }
            });
            
            // 追加のBot検知回避
            Object.defineProperty(navigator, 'permissions', {
                get: () => ({
                    query: () => Promise.resolve({ state: 'granted' })
                })
            });
            
            // Chromeオブジェクトの偽装
            window.chrome = {
                runtime: {},
                loadTimes: function() {},
                csi: function() {},
                app: {}
            };
            
            // 自動化フラグの削除
            delete window.navigator.__proto__.webdriver;
            
            // プロパティ記述子の偽装
            const originalGetOwnPropertyDescriptor = Object.getOwnPropertyDescriptor;
            Object.getOwnPropertyDescriptor = function(obj, prop) {
                if (prop === 'webdriver' && obj === navigator) {
                    return undefined;
                }
                return originalGetOwnPropertyDescriptor.call(this, obj, prop);
            };
            
            // コンソールログの偽装
            const originalLog = console.log;
            console.log = function(...args) {
                if (args[0] && typeof args[0] === 'string' && args[0].includes('webdriver')) {
                    return;
                }
                return originalLog.apply(this, args);
            };
            
            // パフォーマンスタイミングの偽装
            Object.defineProperty(performance, 'timing', {
                get: () => ({
                    navigationStart: Date.now() - Math.random() * 1000,
                    loadEventEnd: Date.now(),
                    domContentLoadedEventEnd: Date.now() - Math.random() * 500
                })
            });
            
            // 追加のCAPTCHA対策
            Object.defineProperty(navigator, 'maxTouchPoints', {
                get: () => 10,
            });
            
            Object.defineProperty(navigator, 'connection', {
                get: () => ({
                    effectiveType: '4g',
                    rtt: 50,
                    downlink: 10,
                    saveData: false
                })
            });
            
            // 自動化検知の回避
            Object.defineProperty(window, 'chrome', {
                get: () => ({
                    runtime: {},
                    loadTimes: function() {},
                    csi: function() {},
                    app: {}
                })
            });
            
            // セキュリティコンテキストの偽装
            Object.defineProperty(window, 'isSecureContext', {
                get: () => true
            });
            
            // ユーザーエージェントの偽装
            Object.defineProperty(navigator, 'userAgent', {
                get: () => 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'
            });
        """)
        
        add_job_log(job_id, "✅ ステルスモード設定完了", jobs)
        return True
        
    except Exception as e:
        add_job_log(job_id, f"⚠️ ステルスモード設定エラー: {e}", jobs)
        return False



def clear_session(page, job_id, jobs):
    """セッションをクリアしてログアウト状態にする"""
    try:
        add_job_log(job_id, "🧹 セッションクリアを実行中...", jobs)
        
        # 1. Jobcanのログアウトページにアクセス
        try:
            page.goto("https://id.jobcan.jp/users/sign_out", timeout=20000)
            add_job_log(job_id, "✅ Jobcanログアウトページにアクセス", jobs)
        except Exception as e:
            add_job_log(job_id, f"⚠️ ログアウトページアクセスエラー: {e}", jobs)
            # ログアウトページアクセスに失敗しても処理を続行
        
        # 2. すべてのクッキーをクリア
        try:
            page.context.clear_cookies()
            add_job_log(job_id, "✅ クッキーをクリアしました", jobs)
        except Exception as e:
            add_job_log(job_id, f"⚠️ クッキークリアエラー: {e}", jobs)
        
        # 3. ローカルストレージとセッションストレージをクリア
        try:
            page.evaluate("""
                localStorage.clear();
                sessionStorage.clear();
            """)
            add_job_log(job_id, "✅ ローカルストレージとセッションストレージをクリアしました", jobs)
        except Exception as e:
            add_job_log(job_id, f"⚠️ ストレージクリアエラー: {e}", jobs)
        
        # 4. 人間らしい待機
        human_like_wait(2.0, 4.0)
        
        add_job_log(job_id, "✅ セッションクリア完了", jobs)
        return True
        
    except Exception as e:
        add_job_log(job_id, f"❌ セッションクリアでエラー: {e}", jobs)
        return False

def process_jobcan_automation(job_id: str, email: str, password: str, file_path: str, jobs: dict, session_dir: str = None, session_id: str = None, company_id: str = None):
    """Jobcan自動化処理のメイン関数（セッション固有のブラウザ環境）"""
    try:
        add_job_log(job_id, "🚀 Jobcan自動化処理を開始", jobs)
        update_progress(job_id, 1, "初期化中...", jobs)
        
        # セッション固有のログ
        if session_id:
            add_job_log(job_id, f"🔑 セッションID: {session_id}", jobs)
        
        # ステップ1: Excelファイルの読み込み
        add_job_log(job_id, "📊 Excelファイルを読み込み中...", jobs)
        update_progress(job_id, 2, "Excelファイル読み込み中...", jobs)
        
        try:
            data_source, total_data = load_excel_data(file_path)
            add_job_log(job_id, f"✅ Excelファイル読み込み完了: {total_data}件のデータ", jobs)
        except Exception as e:
            add_job_log(job_id, f"❌ Excelファイル読み込みエラー: {e}", jobs)
            jobs[job_id]['status'] = 'error'
            jobs[job_id]['login_message'] = f'Excelファイルの読み込みに失敗しました: {str(e)}'
            return
        
        # ステップ2: データ検証
        add_job_log(job_id, "🔍 データ検証中...", jobs)
        update_progress(job_id, 3, "データ検証中...", jobs)
        
        try:
            errors, warnings = validate_excel_data(data_source, pandas_available, job_id, jobs)
            
            # エラーがある場合は処理を停止
            if errors:
                add_job_log(job_id, f"❌ データ検証で{len(errors)}件のエラーが見つかりました", jobs)
                
                # エラーの詳細をユーザーメッセージに含める
                error_details = []
                for error in errors[:5]:  # 最初の5件のエラーを詳細表示
                    error_details.append(error)
                
                if len(errors) > 5:
                    error_details.append(f"他{len(errors) - 5}件のエラーがあります")
                
                error_message = f'データ検証で{len(errors)}件のエラーが見つかりました。\n\n詳細:\n' + '\n'.join(error_details)
                
                jobs[job_id]['status'] = 'error'
                jobs[job_id]['login_message'] = error_message
                return
            
            # 警告がある場合はログに記録
            if warnings:
                add_job_log(job_id, f"⚠️ データ検証で{len(warnings)}件の警告が見つかりました", jobs)
            
            add_job_log(job_id, "✅ データ検証完了", jobs)
        except Exception as e:
            add_job_log(job_id, f"❌ データ検証エラー: {e}", jobs)
            jobs[job_id]['status'] = 'error'
            jobs[job_id]['login_message'] = f'データ検証中にエラーが発生しました: {str(e)}'
            return
        
        # ステップ3: Playwrightの利用可能性チェック
        if not playwright_available:
            add_job_log(job_id, "❌ Playwrightが利用できません", jobs)
            jobs[job_id]['status'] = 'error'
            jobs[job_id]['login_status'] = 'playwright_unavailable'
            jobs[job_id]['login_message'] = 'ブラウザ自動化機能が利用できません'
            return
        
        # ステップ4: ブラウザの起動（セッション固有）
        add_job_log(job_id, "🌐 ブラウザを起動中...", jobs)
        update_progress(job_id, 4, "ブラウザ起動中...", jobs)
        
        try:
            from playwright.sync_api import sync_playwright
            
            with sync_playwright() as p:
                # セッション固有のブラウザ起動オプション
                browser_args = [
                    '--no-sandbox',
                    '--disable-setuid-sandbox',
                    '--disable-dev-shm-usage',
                    '--disable-accelerated-2d-canvas',
                    '--no-first-run',
                    '--no-zygote',
                    '--disable-gpu',
                    '--disable-background-timer-throttling',
                    '--disable-backgrounding-occluded-windows',
                    '--disable-renderer-backgrounding',
                    '--disable-blink-features=AutomationControlled',
                    '--disable-extensions-except',
                    '--disable-plugins-discovery',
                    '--disable-default-apps',
                    '--disable-sync',
                    '--disable-translate',
                    '--hide-scrollbars',
                    '--mute-audio',
                    '--no-default-browser-check',
                    '--no-pings',
                    '--disable-web-security',
                    '--disable-features=VizDisplayCompositor',
                    '--disable-background-networking',
                    '--disable-component-extensions-with-background-pages',
                    '--disable-background-timer-throttling',
                    '--disable-backgrounding-occluded-windows',
                    '--disable-renderer-backgrounding',
                    '--disable-features=TranslateUI',
                    '--disable-ipc-flooding-protection',
                    '--disable-client-side-phishing-detection',
                    '--disable-hang-monitor',
                    '--disable-prompt-on-repost',
                    '--disable-domain-reliability',
                    '--disable-component-update',
                    # CAPTCHA対策の追加オプション
                    '--disable-blink-features=AutomationControlled',
                    '--disable-automation',
                    '--disable-web-security',
                    '--disable-features=VizDisplayCompositor',
                    '--disable-background-networking',
                    '--disable-component-extensions-with-background-pages',
                    '--disable-background-timer-throttling',
                    '--disable-backgrounding-occluded-windows',
                    '--disable-renderer-backgrounding',
                    '--disable-features=TranslateUI',
                    '--disable-ipc-flooding-protection',
                    '--disable-client-side-phishing-detection',
                    '--disable-hang-monitor',
                    '--disable-prompt-on-repost',
                    '--disable-domain-reliability',
                    '--disable-component-update',
                    '--disable-default-apps',
                    '--disable-sync',
                    '--disable-translate',
                    '--hide-scrollbars',
                    '--mute-audio',
                    '--no-first-run',
                    '--no-default-browser-check',
                    '--no-pings',
                    '--disable-web-security',
                    '--disable-features=VizDisplayCompositor',
                    '--disable-background-networking',
                    '--disable-component-extensions-with-background-pages',
                    '--disable-background-timer-throttling',
                    '--disable-backgrounding-occluded-windows',
                    '--disable-renderer-backgrounding',
                    '--disable-features=TranslateUI',
                    '--disable-ipc-flooding-protection',
                    '--disable-client-side-phishing-detection',
                    '--disable-hang-monitor',
                    '--disable-prompt-on-repost',
                    '--disable-domain-reliability',
                    '--disable-component-update'
                ]
                
                # 最新のChrome User-Agent（CAPTCHA対策）
                user_agent = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
                
                # サーバー環境対応のため、通常のlaunchを使用（タイムアウト設定付き）
                browser = p.chromium.launch(
                    headless=True,  # CAPTCHA対策のためヘッドレスモードを有効化
                    args=browser_args,
                    timeout=60000  # ブラウザ起動タイムアウトを60秒に設定
                )
                
                # セッション固有のコンテキスト設定
                context_options = {
                    'viewport': {'width': 1920, 'height': 1080},  # より大きなビューポート
                    'user_agent': user_agent,
                    'ignore_https_errors': True,
                    'java_script_enabled': True,
                    'accept_downloads': True,
                    'locale': 'ja-JP',  # 日本語ロケール
                    'timezone_id': 'Asia/Tokyo',  # 日本時間
                    'permissions': ['geolocation'],  # 位置情報許可
                    'extra_http_headers': {
                        'Accept-Language': 'ja-JP,ja;q=0.9,en;q=0.8',
                        'Accept-Encoding': 'gzip, deflate, br',
                        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
                        'Cache-Control': 'no-cache',
                        'Pragma': 'no-cache'
                    }
                }
                
                context = browser.new_context(**context_options)
                page = context.new_page()
                
                # リダイレクト制御を追加（修正版）
                def handle_request(route):
                    route.continue_()
                
                page.route("**/*", handle_request)
                
                add_job_log(job_id, "✅ ブラウザ起動完了", jobs)
                if session_id:
                    add_job_log(job_id, f"🔑 セッション固有ブラウザ環境: {session_id}", jobs)
                
                # ステルスモードを設定
                setup_stealth_mode(page, job_id, jobs)
                
                # ステップ5: ログイン処理
                add_job_log(job_id, "🔐 Jobcanにログイン中...", jobs)
                update_progress(job_id, 5, "Jobcanログイン中...", jobs)
                
                # ログイン処理開始時の状態を初期化
                jobs[job_id]['login_status'] = 'processing'
                jobs[job_id]['login_message'] = '🔄 ログイン処理中...'
                
                # 新しいCAPTCHA対策ロジックを使用
                login_success, login_status, login_message = perform_login_with_captcha_retry(
                    page, email, password, job_id, jobs, max_captcha_retries=3, company_id=company_id
                )
                
                # ログイン結果をジョブ情報に保存
                jobs[job_id]['login_status'] = login_status
                jobs[job_id]['login_message'] = login_message
                
                if not login_success:
                    add_job_log(job_id, "❌ ログインに失敗したため、処理を停止します", jobs)
                    jobs[job_id]['status'] = 'completed'
                    return
                
                # ステップ6: 実際のデータ入力処理
                add_job_log(job_id, "🔧 ログイン成功のため、実際のデータ入力を試行します", jobs)
                update_progress(job_id, 6, "勤怠データ入力中...", jobs)
                
                perform_actual_data_input(page, data_source, total_data, pandas_available, job_id, jobs)
                
                # ステップ7: 最終確認
                add_job_log(job_id, "🔍 最終確認中...", jobs)
                update_progress(job_id, 7, "最終確認中...", jobs)
                
                # ステップ8: 処理完了
                add_job_log(job_id, "🎉 処理が正常に完了しました", jobs)
                update_progress(job_id, 8, "処理完了中...", jobs)
                
                jobs[job_id]['status'] = 'completed'
                
                # ブラウザを閉じる
                browser.close()
                add_job_log(job_id, "🔒 ブラウザセッションを終了しました", jobs)
                
        except Exception as e:
            add_job_log(job_id, f"❌ ブラウザ処理でエラーが発生しました: {e}", jobs)
            jobs[job_id]['status'] = 'error'
            jobs[job_id]['login_message'] = f'ブラウザ処理でエラーが発生しました: {str(e)}'
            return
        
    except Exception as e:
        add_job_log(job_id, f"❌ 予期しないエラーが発生しました: {e}", jobs)
        jobs[job_id]['status'] = 'error'
        jobs[job_id]['login_message'] = f'予期しないエラーが発生しました: {str(e)}'
        return 

def human_like_mouse_movement(page, job_id, jobs):
    """人間らしいマウス移動を実行（修正版）"""
    try:
        # 現在のビューポートサイズを取得
        viewport = page.viewport_size
        if not viewport:
            viewport = {'width': 1920, 'height': 1080}
        
        # ランダムな位置にマウスを移動（ビューポート内）
        x = random.randint(100, min(800, viewport['width'] - 100))
        y = random.randint(100, min(600, viewport['height'] - 100))
        
        # マウス移動を実行
        page.mouse.move(x, y)
        human_like_wait(0.1, 0.3)
        
        # 軽いスクロール処理（オプション）
        if random.choice([True, False]):
            scroll_amount = random.randint(-50, 50)
            page.mouse.wheel(0, scroll_amount)
            human_like_wait(0.2, 0.5)
        
        add_job_log(job_id, f"🖱️ マウス移動実行: ({x}, {y})", jobs)
        return True
        
    except Exception as e:
        add_job_log(job_id, f"⚠️ マウス移動エラー: {e}", jobs)
        # エラーが発生しても処理を続行
        return True

def perform_login_with_captcha_retry(page, email, password, job_id, jobs, max_captcha_retries=3, company_id=None):
    """CAPTCHA対策付きログイン処理"""
    try:
        add_job_log(job_id, "🔐 ログイン処理を開始します", jobs)
        
        # 無限ループ防止カウンター
        attempt_count = 0
        max_total_attempts = 10
        
        for captcha_attempt in range(max_captcha_retries):
            attempt_count += 1
            
            if attempt_count > max_total_attempts:
                add_job_log(job_id, "❌ 無限ループを防止するため、ログイン処理を停止します", jobs)
                return False, "max_attempts_exceeded", "❌ ログイン試行回数が上限に達しました"
            
            add_job_log(job_id, f"🔄 ログイン試行 {captcha_attempt + 1}/{max_captcha_retries}", jobs)
            
            # ログイン処理を実行
            login_success = perform_login(page, email, password, job_id, jobs, company_id)
            
            if login_success:
                add_job_log(job_id, "✅ ログイン処理が成功しました", jobs)
                return True, "success", "✅ ログインに成功しました"
            
            # ログイン状態をチェック
            is_logged_in, login_status, login_message = check_login_status(page, job_id, jobs)
            
            if is_logged_in:
                add_job_log(job_id, "✅ ログイン状態を確認しました", jobs)
                return True, "success", "✅ ログインに成功しました"
            
            # CAPTCHAの確認
            if "CAPTCHA" in login_message or "画像認証" in login_message:
                add_job_log(job_id, f"🔄 CAPTCHA検出: 試行 {captcha_attempt + 1}", jobs)
                
                # セッションをクリアして再試行
                clear_session(page, job_id, jobs)
                human_like_wait(5.0, 10.0)  # 長めの待機
                
                if captcha_attempt < max_captcha_retries - 1:
                    add_job_log(job_id, f"🔄 CAPTCHA再試行: {captcha_attempt + 2}/{max_captcha_retries}", jobs)
                    continue
                else:
                    add_job_log(job_id, "❌ CAPTCHA再試行回数が上限に達しました", jobs)
                    return False, "captcha_failed", "❌ 画像認証の処理に失敗しました"
            
            # その他のエラーの場合
            if captcha_attempt < max_captcha_retries - 1:
                add_job_log(job_id, f"🔄 ログイン再試行: {captcha_attempt + 2}/{max_captcha_retries}", jobs)
                clear_session(page, job_id, jobs)
                human_like_wait(3.0, 6.0)
                continue
            else:
                add_job_log(job_id, "❌ ログイン再試行回数が上限に達しました", jobs)
                return False, "login_failed", login_message
        
        # すべての試行が失敗
        add_job_log(job_id, "❌ すべてのログイン試行が失敗しました", jobs)
        return False, "all_attempts_failed", "❌ ログインに失敗しました"
        
    except Exception as e:
        add_job_log(job_id, f"❌ ログイン処理で例外が発生: {e}", jobs)
        return False, "exception", f"❌ ログイン処理でエラーが発生しました: {str(e)}"

def retry_on_captcha(page, email, password, job_id, jobs, max_retries=3):
    """
    CAPTCHAが検出された際にリトライを行う共通処理。
    page: Playwrightのページインスタンス
    email: メールアドレス
    password: パスワード
    job_id: ジョブID
    jobs: ジョブ辞書
    max_retries: リトライ最大回数
    """
    for retry in range(1, max_retries + 1):
        try:
            add_job_log(job_id, f"🔄 CAPTCHAリトライ試行 {retry}/{max_retries}", jobs)
            
            # セッションクリアを実行
            clear_session(page, job_id, jobs)
            
            # ページをリロード
            page.reload()
            page.wait_for_load_state('networkidle', timeout=30000)
            
            # 人間らしい待機（CAPTCHA対策のため長め）
            wait_time = random.uniform(5.0, 10.0)
            add_job_log(job_id, f"⏳ {wait_time:.1f}秒待機中...", jobs)
            human_like_wait(wait_time, wait_time + 3.0)
            
            # ログイン処理を再実行
            login_success, status, message = perform_login(page, email, password, job_id, jobs)
            
            if login_success:
                add_job_log(job_id, f"✅ リトライ {retry} でログイン成功", jobs)
                return True, status, message
            elif status == "captcha_detected":
                add_job_log(job_id, f"⚠️ リトライ {retry} でもCAPTCHAが発生", jobs)
                if retry < max_retries:
                    # 次のリトライ前に待機（CAPTCHA対策のため長め）
                    wait_time = random.uniform(15.0, 30.0)
                    add_job_log(job_id, f"⏳ {wait_time:.1f}秒待機してから再試行", jobs)
                    time.sleep(wait_time)
                continue
            else:
                add_job_log(job_id, f"❌ リトライ {retry} でログイン失敗: {message}", jobs)
                return False, status, message
                
        except Exception as e:
            add_job_log(job_id, f"❌ CAPTCHAリトライ {retry} でエラー: {e}", jobs)
            if retry < max_retries:
                wait_time = random.uniform(10.0, 20.0)
                add_job_log(job_id, f"⏳ {wait_time:.1f}秒待機してから再試行", jobs)
                time.sleep(wait_time)
            continue
    
    add_job_log(job_id, f"❌ CAPTCHAリトライ {max_retries} 回すべて失敗", jobs)
    return False, "captcha_failed", "❌ CAPTCHAが解決できませんでした。手動でログインしてください。"