import time
import tempfile
import numbers
from datetime import datetime, date
import calendar
import re
import os
import uuid
import importlib.util
from typing import Tuple, List, Optional

# ライブラリの利用可能性をチェック
try:
    import pandas as pd
    pandas_available = True
except ImportError:
    pandas_available = False

try:
    import openpyxl
    from openpyxl import Workbook
    openpyxl_available = True
except ImportError:
    openpyxl_available = False

try:
    import jpholiday
    jpholiday_available = True
except ImportError:
    jpholiday_available = False

playwright_available = importlib.util.find_spec("playwright") is not None

def get_weekdays_in_current_month():
    """本日の月の平日（土日・祝日を除く）を取得"""
    today = date.today()
    year = today.year
    month = today.month
    
    # その月の1日から末日までを取得
    _, last_day = calendar.monthrange(year, month)
    weekdays = []
    
    for day in range(1, last_day + 1):
        current_date = date(year, month, day)
        
        # 土日チェック（0=月曜日, 6=日曜日）
        if current_date.weekday() >= 5:  # 土曜日(5)または日曜日(6)
            continue
        
        # 祝日チェック
        if jpholiday_available:
            if jpholiday.is_holiday(current_date):
                continue
        
        # 平日の場合、YYYY-MM-DD形式で追加
        weekdays.append(current_date.strftime('%Y-%m-%d'))
    
    return weekdays

def get_weekdays_in_previous_month():
    """先月の平日（土日・祝日を除く）を取得"""
    today = date.today()
    
    # 先月の年月を計算
    if today.month == 1:
        prev_year = today.year - 1
        prev_month = 12
    else:
        prev_year = today.year
        prev_month = today.month - 1
    
    # 先月の1日から末日までを取得
    _, last_day = calendar.monthrange(prev_year, prev_month)
    weekdays = []
    
    for day in range(1, last_day + 1):
        current_date = date(prev_year, prev_month, day)
        
        # 土日チェック（0=月曜日, 6=日曜日）
        if current_date.weekday() >= 5:  # 土曜日(5)または日曜日(6)
            continue
        
        # 祝日チェック
        if jpholiday_available:
            if jpholiday.is_holiday(current_date):
                continue
        
        # 平日の場合、YYYY-MM-DD形式で追加
        weekdays.append(current_date.strftime('%Y-%m-%d'))
    
    return weekdays

def validate_excel_data(data_source, pandas_available, job_id, jobs):
    """Excelデータの内容を検証（空白行スキップ対応）"""
    errors = []
    warnings = []
    
    try:
        if pandas_available:
            # pandasを使用した検証
            if len(data_source) == 0:
                errors.append("Excelファイルにデータが含まれていません")
                return errors, warnings
            
            # ヘッダー行の確認
            expected_columns = ['日付', '開始時刻', '終了時刻']
            actual_columns = list(data_source.columns)
            
            if not all(col in actual_columns for col in expected_columns):
                missing_columns = [col for col in expected_columns if col not in actual_columns]
                errors.append(f"必要な列が見つかりません。不足している列: {missing_columns}")
                return errors, warnings
            
            # 空白行をスキップするためのフィルタ処理
            add_job_log(job_id, "🔍 空白行のスキップ処理を実行中...", jobs)
            
            # 日付列が空の行を除外
            filtered_data = data_source.dropna(subset=['日付'], how='all')
            
            # 主要カラムがすべて空の行も除外
            key_columns = ['日付', '開始時刻', '終了時刻']
            filtered_data = filtered_data.dropna(subset=key_columns, how='all')
            
            # スキップされた行数を計算
            skipped_rows = len(data_source) - len(filtered_data)
            if skipped_rows > 0:
                add_job_log(job_id, f"✅ 空白行 {skipped_rows} 行をスキップしました", jobs)
            
            # フィルタ後のデータ件数をログに記録
            total_rows = len(filtered_data)
            add_job_log(job_id, f"✅ Excelファイル読み込み完了: {total_rows}件の有効データ", jobs)
            
            # フィルタ後のデータで各行を検証
            for index, row in filtered_data.iterrows():
                row_num = index + 2  # ヘッダー行を考慮
                
                # 日付の検証（空白行は既にスキップ済み）
                date_value = row.iloc[0]
                if pd.isna(date_value):
                    # この時点で空白行がある場合は、データ構造の問題
                    errors.append(f"{row_num}行目の「日付」が空白です（データ構造エラー）")
                    continue
                
                # 日付形式の検証
                try:
                    if isinstance(date_value, str):
                        # 文字列の場合、複数の形式を試行
                        date_formats = ['%Y-%m-%d', '%Y/%m/%d', '%Y年%m月%d日']
                        parsed_date = None
                        for fmt in date_formats:
                            try:
                                parsed_date = datetime.strptime(date_value, fmt).date()
                                break
                            except ValueError:
                                continue
                        
                        if parsed_date is None:
                            errors.append(f"{row_num}行目の「日付」の形式が無効です: {date_value} (期待形式: YYYY-MM-DD, YYYY/MM/DD, YYYY年MM月DD日)")
                            continue
                    else:
                        parsed_date = pd.to_datetime(date_value).date()
                    
                    # 未来日チェック
                    if parsed_date > date.today():
                        warnings.append(f"{row_num}行目の「日付」が未来の日付です: {parsed_date}")
                    
                    # 過去すぎる日付チェック（1年前まで）
                    one_year_ago = date.today().replace(year=date.today().year - 1)
                    if parsed_date < one_year_ago:
                        warnings.append(f"{row_num}行目の「日付」が過去すぎます: {parsed_date}")
                    
                except Exception as e:
                    errors.append(f"{row_num}行目の「日付」の解析に失敗しました: {date_value}")
                    continue
                
                # 時刻の検証
                start_time = row.iloc[1]
                end_time = row.iloc[2]
                
                # 開始時刻の検証と正規化
                normalized_start_time, start_error = validate_time_value(start_time, row_num, "開始時刻")
                if start_error:
                    errors.append(start_error)
                    continue
                
                # 終了時刻の検証と正規化
                normalized_end_time, end_error = validate_time_value(end_time, row_num, "終了時刻")
                if end_error:
                    errors.append(end_error)
                    continue
                
                # 勤務時間の妥当性チェック
                try:
                    # 正規化された時刻を使用
                    start_parts = normalized_start_time.split(':')
                    end_parts = normalized_end_time.split(':')
                    
                    if len(start_parts) >= 2 and len(end_parts) >= 2:
                        start_minutes = int(start_parts[0]) * 60 + int(start_parts[1])
                        end_minutes = int(end_parts[0]) * 60 + int(end_parts[1])
                        
                        if end_minutes <= start_minutes and int(end_parts[0]) < 12 <= int(start_parts[0]):
                            end_minutes += 24 * 60

                        # 勤務時間が短すぎる場合
                        work_hours = (end_minutes - start_minutes) / 60
                        if work_hours < 0.5:  # 30分未満
                            warnings.append(f"{row_num}行目: 勤務時間が短すぎます: {work_hours:.1f}時間")
                        elif work_hours > 24:  # 24時間超過
                            warnings.append(f"{row_num}行目: 勤務時間が長すぎます: {work_hours:.1f}時間")
                        
                        # 開始時刻が終了時刻より後の場合
                        if start_minutes >= end_minutes:
                            errors.append(f"{row_num}行目: 「開始時刻」が「終了時刻」より後です: {normalized_start_time} > {normalized_end_time}")
                    
                except Exception as e:
                    warnings.append(f"{row_num}行目: 勤務時間の計算に失敗しました: {e}")
            
        else:
            # openpyxlを使用した検証
            ws = data_source.active
            if ws.max_row <= 1:
                errors.append("Excelファイルにデータが含まれていません")
                return errors, warnings
            
            # 空白行をスキップするためのフィルタ処理
            add_job_log(job_id, "🔍 空白行のスキップ処理を実行中...", jobs)
            
            # 有効な行のみを収集
            valid_rows = []
            skipped_count = 0
            
            for row in range(2, ws.max_row + 1):
                # 主要カラムの値を取得
                date_value = ws[f'A{row}'].value
                start_time_value = ws[f'B{row}'].value
                end_time_value = ws[f'C{row}'].value
                
                # すべての主要カラムが空の場合はスキップ
                if (date_value is None or str(date_value).strip() == '') and \
                   (start_time_value is None or str(start_time_value).strip() == '') and \
                   (end_time_value is None or str(end_time_value).strip() == ''):
                    skipped_count += 1
                    continue
                
                valid_rows.append(row)
            
            if skipped_count > 0:
                add_job_log(job_id, f"✅ 空白行 {skipped_count} 行をスキップしました", jobs)
            
            # フィルタ後のデータ件数をログに記録
            total_rows = len(valid_rows)
            add_job_log(job_id, f"✅ Excelファイル読み込み完了: {total_rows}件の有効データ", jobs)
            
            # 有効な行のみを検証
            for row in valid_rows:
                row_num = row
                
                # 日付の検証（空白行は既にスキップ済み）
                date_value = ws[f'A{row}'].value
                if date_value is None or str(date_value).strip() == '':
                    # この時点で空白行がある場合は、データ構造の問題
                    errors.append(f"{row_num}行目の「日付」が空白です（データ構造エラー）")
                    continue
                
                # 時刻の検証
                start_time = ws[f'B{row}'].value
                end_time = ws[f'C{row}'].value
                
                # 開始時刻の検証と正規化
                normalized_start_time, start_error = validate_time_value(start_time, row_num, "開始時刻")
                if start_error:
                    errors.append(start_error)
                    continue
                
                # 終了時刻の検証と正規化
                normalized_end_time, end_error = validate_time_value(end_time, row_num, "終了時刻")
                if end_error:
                    errors.append(end_error)
                    continue
                
                # 勤務時間の妥当性チェック
                try:
                    # 正規化された時刻を使用
                    start_parts = normalized_start_time.split(':')
                    end_parts = normalized_end_time.split(':')
                    
                    if len(start_parts) >= 2 and len(end_parts) >= 2:
                        start_minutes = int(start_parts[0]) * 60 + int(start_parts[1])
                        end_minutes = int(end_parts[0]) * 60 + int(end_parts[1])
                        
                        if end_minutes <= start_minutes and int(end_parts[0]) < 12 <= int(start_parts[0]):
                            end_minutes += 24 * 60

                        # 勤務時間が短すぎる場合
                        work_hours = (end_minutes - start_minutes) / 60
                        if work_hours < 0.5:  # 30分未満
                            warnings.append(f"{row_num}行目: 勤務時間が短すぎます: {work_hours:.1f}時間")
                        elif work_hours > 24:  # 24時間超過
                            warnings.append(f"{row_num}行目: 勤務時間が長すぎます: {work_hours:.1f}時間")
                        
                        # 開始時刻が終了時刻より後の場合
                        if start_minutes >= end_minutes:
                            errors.append(f"{row_num}行目: 「開始時刻」が「終了時刻」より後です: {normalized_start_time} > {normalized_end_time}")
                    
                except Exception as e:
                    warnings.append(f"{row_num}行目: 勤務時間の計算に失敗しました: {e}")
        
    except Exception as e:
        errors.append(f"データ検証中にエラーが発生しました: {e}")
    
    # 検証結果をログに記録
    if errors:
        add_job_log(job_id, f"❌ データ検証エラー: {len(errors)}件", jobs)
        for error in errors[:10]:  # 最初の10件のみ表示
            add_job_log(job_id, f"  - {error}", jobs)
        if len(errors) > 10:
            add_job_log(job_id, f"  - 他{len(errors) - 10}件のエラーがあります", jobs)
    
    if warnings:
        add_job_log(job_id, f"⚠️ データ検証警告: {len(warnings)}件", jobs)
        for warning in warnings[:5]:  # 最初の5件のみ表示
            add_job_log(job_id, f"  - {warning}", jobs)
        if len(warnings) > 5:
            add_job_log(job_id, f"  - 他{len(warnings) - 5}件の警告があります", jobs)
    
    return errors, warnings

def allowed_file(filename):
    """アップロードされたファイルの拡張子をチェック"""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in {'xlsx', 'xls'}

# P1: ログの上限設定（メモリ最適化）。監査対応で500に短縮し単調増加を抑制
MAX_JOB_LOGS = 500  # 最大ログ件数（1ジョブあたり）
MAX_LOG_CHARS = 2000  # 1ログエントリの最大文字数
MAX_JOB_LOG_BYTES = 200 * 1024  # 最大ログサイズ（200KB、バイト単位）

def add_job_log(job_id: str, message: str, jobs: dict):
    """
    ジョブのログを追加（P1: メモリ上限設定付き）
    
    Args:
        job_id: ジョブID
        message: ログメッセージ
        jobs: ジョブ辞書
    
    Note:
        - collections.deque(maxlen=MAX_JOB_LOGS)を使用して固定長ログを実現
        - ログサイズがMAX_JOB_LOG_BYTESを超えた場合、古いログから削除
        - JSON返却時はlist()に変換して互換性を維持
    """
    if job_id not in jobs:
        return
    
    from collections import deque
    
    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    sanitized_message = sanitize_log_message(message)
    
    # P1: メッセージ長制限
    if len(sanitized_message) > MAX_LOG_CHARS:
        sanitized_message = sanitized_message[:MAX_LOG_CHARS - 3] + "..."
    
    log_entry = f"[{timestamp}] {sanitized_message}"
    
    # P1: dequeを使用して固定長ログを実現（効率的なメモリ管理）
    if 'logs' not in jobs[job_id]:
        jobs[job_id]['logs'] = deque(maxlen=MAX_JOB_LOGS)
    
    # dequeでない場合は変換（既存データの互換性）
    if not isinstance(jobs[job_id]['logs'], deque):
        jobs[job_id]['logs'] = deque(jobs[job_id]['logs'], maxlen=MAX_JOB_LOGS)
    
    # ログを追加（maxlenにより自動的に古いログが削除される）
    jobs[job_id]['logs'].append(log_entry)
    
    # P1: ログサイズ制限（バイト単位）
    # ログサイズが上限を超えた場合、古いログから削除
    total_bytes = sum(len(log.encode('utf-8')) for log in jobs[job_id]['logs'])
    if total_bytes > MAX_JOB_LOG_BYTES:
        # 古いログから削除（末尾のみ保持）
        logs_list = list(jobs[job_id]['logs'])
        current_bytes = 0
        keep_logs = []
        
        # 末尾から逆順に追加（新しいログを優先）
        for log in reversed(logs_list):
            log_bytes = len(log.encode('utf-8'))
            if current_bytes + log_bytes <= MAX_JOB_LOG_BYTES:
                keep_logs.insert(0, log)
                current_bytes += log_bytes
            else:
                break
        
        jobs[job_id]['logs'] = deque(keep_logs, maxlen=MAX_JOB_LOGS)

def sanitize_log_message(message):
    """ログメッセージから個人情報を除去"""
    # メールアドレスを除去
    message = re.sub(r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b', '[EMAIL]', message)
    
    # パスワードを除去
    message = re.sub(r'password["\']?\s*[:=]\s*["\']?[^"\s]+["\']?', 'password="[PASSWORD]"', message, flags=re.IGNORECASE)
    
    # クレジットカード番号を除去
    message = re.sub(r'\b\d{4}[- ]?\d{4}[- ]?\d{4}[- ]?\d{4}\b', '[CARD_NUMBER]', message)
    
    return message

def update_progress(job_id: str, step: int, step_name: str, jobs: dict, current_data: int = 0, total_data: int = 0):
    """ジョブの進捗を更新"""
    if job_id in jobs:
        jobs[job_id]['step'] = step
        jobs[job_id]['step_name'] = step_name
        jobs[job_id]['current_data'] = current_data
        jobs[job_id]['total_data'] = total_data

def create_template_excel():
    """今月の平日のテンプレートExcelファイルを作成"""
    try:
        print("テンプレート作成開始")
        
        # 今月の平日を取得
        weekdays = get_weekdays_in_current_month()
        print(f"平日数: {len(weekdays)}")
        
        if not weekdays:
            print("平日が見つかりませんでした")
            return None, "今月の平日が見つかりませんでした"
        
        # サンプルデータを作成
        sample_data = []
        for weekday in weekdays:
            # 9:00-18:00の勤務時間をサンプルとして設定
            sample_data.append({
                '日付': weekday,
                '開始時刻': '09:00',
                '終了時刻': '18:00'
            })
        
        print(f"サンプルデータ作成完了: {len(sample_data)}件")
        
        # テンプレートファイルを作成
        if openpyxl_available:
            print("openpyxlを使用してテンプレート作成")
            wb = Workbook()
            ws = wb.active
            ws.title = "勤怠データ"
            
            # ヘッダー行を追加
            ws['A1'] = '日付'
            ws['B1'] = '開始時刻'
            ws['C1'] = '終了時刻'
            
            # データ行を追加
            for i, data in enumerate(sample_data, start=2):
                ws[f'A{i}'] = data['日付']
                ws[f'B{i}'] = data['開始時刻']
                ws[f'C{i}'] = data['終了時刻']
            
            # 本番環境に対応した一時ファイルの生成
            # /tmpディレクトリを使用（Renderなどの本番環境で安全）
            temp_dir = '/tmp' if os.path.exists('/tmp') else tempfile.gettempdir()
            print(f"一時ディレクトリ: {temp_dir}")
            
            temp_file = tempfile.NamedTemporaryFile(
                delete=False, 
                suffix='.xlsx',
                dir=temp_dir
            )
            
            try:
                print(f"ファイル保存開始: {temp_file.name}")
                wb.save(temp_file.name)
                temp_file.close()
                
                # ファイルが正常に作成されたか確認
                if os.path.exists(temp_file.name) and os.path.getsize(temp_file.name) > 0:
                    file_size = os.path.getsize(temp_file.name)
                    print(f"テンプレートファイル作成成功: {temp_file.name} ({file_size} bytes)")
                    return temp_file.name, None
                else:
                    print("テンプレートファイルの保存に失敗しました")
                    return None, "テンプレートファイルの保存に失敗しました"
                    
            except Exception as save_error:
                print(f"ファイル保存エラー: {save_error}")
                # 一時ファイルのクリーンアップ
                try:
                    if os.path.exists(temp_file.name):
                        os.remove(temp_file.name)
                except:
                    pass
                return None, f"ファイル保存エラー: {str(save_error)}"
        else:
            print("openpyxlが利用できません")
            return None, "openpyxlが利用できません"
    
    except Exception as e:
        print(f"テンプレート作成例外: {e}")
        return None, f"テンプレート作成中にエラーが発生しました: {e}"

def create_previous_month_template_excel():
    """先月の平日のテンプレートExcelファイルを作成"""
    try:
        print("先月テンプレート作成開始")
        
        # 先月の平日を取得
        weekdays = get_weekdays_in_previous_month()
        print(f"先月平日数: {len(weekdays)}")
        
        if not weekdays:
            print("先月の平日が見つかりませんでした")
            return None, "先月の平日が見つかりませんでした"
        
        # サンプルデータを作成
        sample_data = []
        for weekday in weekdays:
            # 9:00-18:00の勤務時間をサンプルとして設定
            sample_data.append({
                '日付': weekday,
                '開始時刻': '09:00',
                '終了時刻': '18:00'
            })
        
        print(f"先月サンプルデータ作成完了: {len(sample_data)}件")
        
        # テンプレートファイルを作成
        if openpyxl_available:
            print("openpyxlを使用して先月テンプレート作成")
            wb = Workbook()
            ws = wb.active
            ws.title = "勤怠データ"
            
            # ヘッダー行を追加
            ws['A1'] = '日付'
            ws['B1'] = '開始時刻'
            ws['C1'] = '終了時刻'
            
            # データ行を追加
            for i, data in enumerate(sample_data, start=2):
                ws[f'A{i}'] = data['日付']
                ws[f'B{i}'] = data['開始時刻']
                ws[f'C{i}'] = data['終了時刻']
            
            # 本番環境に対応した一時ファイルの生成
            temp_dir = '/tmp' if os.path.exists('/tmp') else tempfile.gettempdir()
            print(f"一時ディレクトリ: {temp_dir}")
            
            temp_file = tempfile.NamedTemporaryFile(
                delete=False, 
                suffix='.xlsx',
                dir=temp_dir
            )
            
            try:
                print(f"先月ファイル保存開始: {temp_file.name}")
                wb.save(temp_file.name)
                temp_file.close()
                
                # ファイルが正常に作成されたか確認
                if os.path.exists(temp_file.name) and os.path.getsize(temp_file.name) > 0:
                    file_size = os.path.getsize(temp_file.name)
                    print(f"先月テンプレートファイル作成成功: {temp_file.name} ({file_size} bytes)")
                    return temp_file.name, None
                else:
                    print("先月テンプレートファイルの保存に失敗しました")
                    return None, "先月テンプレートファイルの保存に失敗しました"
                    
            except Exception as save_error:
                print(f"先月ファイル保存エラー: {save_error}")
                # 一時ファイルのクリーンアップ
                try:
                    if os.path.exists(temp_file.name):
                        os.remove(temp_file.name)
                except:
                    pass
                return None, f"先月ファイル保存エラー: {str(save_error)}"
        else:
            print("openpyxlが利用できません")
            return None, "openpyxlが利用できません"
    
    except Exception as e:
        print(f"先月テンプレート作成例外: {e}")
        return None, f"先月テンプレート作成中にエラーが発生しました: {e}"

def load_excel_data(file_path):
    """Excelファイルを読み込み"""
    try:
        if pandas_available:
            # pandasを使用して読み込み
            import pandas as pd
            data = pd.read_excel(file_path)
            # 有効なデータ行数をカウント（ヘッダー行を除く、空白行も除外）
            valid_rows = data.dropna(subset=['日付'], how='all').dropna(subset=['日付', '開始時刻', '終了時刻'], how='all')
            return data, len(valid_rows)
        elif openpyxl_available:
            # openpyxlを使用して読み込み
            from openpyxl import load_workbook
            wb = load_workbook(file_path)
            ws = wb.active
            
            # 有効なデータ行数をカウント（ヘッダー行を除く、空白行も除外）
            valid_rows = 0
            for row in range(2, ws.max_row + 1):
                date_value = ws[f'A{row}'].value
                start_time_value = ws[f'B{row}'].value
                end_time_value = ws[f'C{row}'].value
                
                # すべての主要カラムが空の場合はスキップ
                if (date_value is None or str(date_value).strip() == '') and \
                   (start_time_value is None or str(start_time_value).strip() == '') and \
                   (end_time_value is None or str(end_time_value).strip() == ''):
                    continue
                
                valid_rows += 1
            
            return wb, valid_rows
        else:
            return None, 0
    except Exception as e:
        print(f"Excelファイル読み込みエラー: {e}")
        return None, 0

def extract_date_info(date):
    """日付から年月日を抽出"""
    if hasattr(date, 'strftime'):
        date_str = date.strftime('%Y-%m-%d')
    else:
        date_str = str(date)
    
    if hasattr(date, 'year') and hasattr(date, 'month') and hasattr(date, 'day'):
        year = date.year
        month = date.month
        day = date.day
    else:
        date_str_clean = str(date_str).strip()
        separators = ['-', '/', '年', '月', '日']
        date_parts = None
        
        for sep in separators:
            if sep in date_str_clean:
                parts = date_str_clean.split(sep)
                if len(parts) >= 3:
                    try:
                        year_part = parts[0].strip()
                        month_part = parts[1].strip()
                        day_part = parts[2].strip()
                        if (year_part.isdigit() and len(year_part) == 4 and
                            month_part.isdigit() and 1 <= int(month_part) <= 12 and
                            day_part.isdigit() and 1 <= int(day_part) <= 31):
                            date_parts = [year_part, month_part, day_part]
                            break
                    except (ValueError, IndexError):
                        continue
        
        if date_parts:
            year = date_parts[0]
            month = date_parts[1]
            day = date_parts[2]
        else:
            year = month = day = "01"
    
    return date_str, year, month, day

def simulate_data_processing(job_id, data_source, total_data, pandas_available, jobs):
    """データ処理のシミュレーション"""
    for i in range(total_data):
        time.sleep(0.1)  # 処理時間をシミュレート
        progress = int((i + 1) / total_data * 100)
        update_progress(job_id, 6, f"データ処理中... ({i + 1}/{total_data})", jobs, i + 1, total_data)
        add_job_log(job_id, f"データ処理進捗: {progress}%", jobs)

def _normalize_time_parts(hours: int, minutes: int) -> Tuple[str, Optional[str]]:
    """HH:MM形式へ正規化する。深夜勤務向けに24時超えも許容する。"""
    if minutes < 0 or minutes > 59:
        return None, f"分が無効です: {minutes}"
    if hours < 0 or hours > 47:
        return None, f"時が無効です: {hours} (0〜47時で入力してください)"
    return f"{hours:02d}:{minutes:02d}", None


def _normalize_from_total_minutes(total_minutes: int) -> Tuple[str, Optional[str]]:
    hours, minutes = divmod(total_minutes, 60)
    return _normalize_time_parts(hours, minutes)


def normalize_time_format(time_value) -> Tuple[str, Optional[str]]:
    """
    時刻値を正規化してHH:MM形式に変換
    
    Args:
        time_value: 時刻値（文字列、datetime.time、その他）
    
    Returns:
        Tuple[str, Optional[str]]: (正規化された時刻文字列, エラーメッセージ)
    """
    try:
        if hasattr(time_value, 'total_seconds'):
            total_minutes = int(round(time_value.total_seconds() / 60))
            return _normalize_from_total_minutes(total_minutes)

        if isinstance(time_value, numbers.Real) and not isinstance(time_value, bool):
            numeric_value = float(time_value)
            if 0 <= numeric_value < 2 and not numeric_value.is_integer():
                total_minutes = int(round(numeric_value * 24 * 60))
                return _normalize_from_total_minutes(total_minutes)
            numeric_str = str(int(numeric_value)) if numeric_value.is_integer() else str(time_value)
            return normalize_time_format(numeric_str)

        if hasattr(time_value, 'hour') and hasattr(time_value, 'minute'):
            return _normalize_time_parts(int(time_value.hour), int(time_value.minute))

        if isinstance(time_value, str):
            time_str = str(time_value).strip()
            if not time_str:
                return None, f"時刻形式が無効です: {time_value}"

            normalized = (
                time_str.replace('：', ':')
                .replace('時', ':')
                .replace('分', '')
                .replace('　', ' ')
                .strip()
            )
            normalized = re.sub(r'\s+', '', normalized)

            match = re.match(r'^(\d{1,2})(?::(\d{1,2}))?(?::\d{1,2}(?:\.\d+)?)?$', normalized)
            if match:
                hours = int(match.group(1))
                minutes = int(match.group(2) or 0)
                return _normalize_time_parts(hours, minutes)

            digits = re.sub(r'\D', '', normalized)
            if digits:
                if len(digits) <= 2:
                    return _normalize_time_parts(int(digits), 0)
                if len(digits) == 3:
                    return _normalize_time_parts(int(digits[0]), int(digits[1:]))
                if len(digits) >= 4:
                    return _normalize_time_parts(int(digits[:-2]), int(digits[-2:]))

            return None, f"時刻形式が無効です: {time_value} (例: 09:00, 900, 26:00, 29時30分)"

        return None, f"時刻形式が無効です: {time_value}"
        
    except Exception as e:
        return None, f"時刻の解析に失敗しました: {time_value} - {str(e)}"

def validate_time_value(time_value, row_num: int, time_name: str) -> Tuple[str, Optional[str]]:
    """
    時刻値を検証して正規化
    
    Args:
        time_value: 時刻値
        row_num: 行番号
        time_name: 時刻名（開始時刻/終了時刻）
    
    Returns:
        Tuple[str, Optional[str]]: (正規化された時刻文字列, エラーメッセージ)
    """
    if time_value is None or (isinstance(time_value, str) and time_value.strip() == ''):
        return None, f"{row_num}行目の「{time_name}」が空白です"
    
    normalized_time, error = normalize_time_format(time_value)
    if error:
        return None, f"{row_num}行目の「{time_name}」{error}"
    
    return normalized_time, None
